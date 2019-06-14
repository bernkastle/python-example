from openpyxl import load_workbook, worksheet
from openpyxl.styles import PatternFill, colors
from collections import Counter


def get_color(color: str) -> str:
    return '9ccc65' if color == '26c6da' else '26c6da'


def excel_change_color_by_value():
    """
    此方法实现了当excel内值变化时，切换行颜色
    即同样的值颜色保持一致，在两种颜色中切换
    :return: None
    """

    current_color = '9ccc65'
    previous = None

    # 打开excel文件
    wb = load_workbook("test.xlsx")
    # 打开excel sheet
    sheet: worksheet = wb["Sheet1"]

    # 便利excel的行
    for irow, row in enumerate(sheet.rows, start=1):
        if row[4].value != previous:
            current_color = get_color(current_color)  # 如果value变化，则切换颜色

        for cell in row:
            fill = PatternFill(start_color=current_color, end_color=current_color, fill_type='solid')  # 设置填充模式和颜色
            cell.fill = fill  # 为excel填充颜色

    # 保存excel
    wb.save('new.xlsx')
